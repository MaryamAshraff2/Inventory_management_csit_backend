from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, timedelta
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from django.conf import settings
import os
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..models import Report, Procurement, StockMovement, Item, SendingStockRequest, DiscardedItem, User, InventoryByLocation, Location, ProcurementItem
from ..serializers import ReportSerializer
# Add Category import
from ..models import Category


class ReportViewSet(viewsets.ModelViewSet):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer

    def _get_generated_by_user(self, request):
        """Helper method to get the appropriate User instance for generated_by"""
        if hasattr(request, 'user') and request.user and not request.user.is_anonymous:
            # Try to find our custom User model instance
            try:
                # If request.user is our custom User model
                if isinstance(request.user, User):
                    return request.user
                # If request.user is Django auth User, try to find corresponding custom User
                elif hasattr(request.user, 'email'):
                    return User.objects.get(email=request.user.email)
                else:
                    return None
            except User.DoesNotExist:
                return None
        return None

    def _parse_date_filter(self, date_str):
        """Helper method to parse date strings and make them timezone-aware"""
        if not date_str:
            return None
        try:
            # Parse the date string
            parsed_date = parse_date(date_str)
            if parsed_date:
                # Make it timezone-aware by combining with current timezone
                return timezone.make_aware(
                    datetime.combine(parsed_date, datetime.min.time()),
                    timezone=timezone.get_current_timezone()
                )
        except (ValueError, TypeError):
            pass
        return None

    def _generate_pdf_content(self, report_data, report_type):
        """Generate professional PDF content based on report type and data"""
        # Special handling for Register report
        if report_type == 'register':
            return self._generate_register_pdf_content(report_data)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=0.75*inch,
            rightMargin=0.75*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        elements = []
        
        # Custom styles
        styles = getSampleStyleSheet()
        
        # Add custom styles
        styles.add(ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#2E5984')
        ))
        
        styles.add(ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#1E3F66')
        ))
        
        styles.add(ParagraphStyle(
            'Metadata',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            fontName='Helvetica'
        ))
        
        styles.add(ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            spaceBefore=20,
            alignment=TA_CENTER,
            textColor=colors.grey
        ))
        
        # Add company logo and header
        try:
            # Try multiple possible paths for the logo
            logo_paths = [
                "nedlogo.jpeg",
                "inventory/views/nedlogo.jpeg",
                os.path.join(os.path.dirname(__file__), "nedlogo.jpeg"),
                os.path.join(settings.BASE_DIR, "inventory", "views", "nedlogo.jpeg")
            ]
            
            logo_found = False
            for logo_path in logo_paths:
                if os.path.exists(logo_path):
                    logo = Image(logo_path, width=1*inch, height=1*inch)
                    elements.append(logo)
                    elements.append(Spacer(1, 10))
                    logo_found = True
                    break
                    
            if not logo_found:
                # Add a text header instead
                header_text = Paragraph("CSIT Inventory Management System", styles['CustomTitle'])
                elements.append(header_text)
                elements.append(Spacer(1, 10))
        except Exception as e:
            # Add a text header instead if logo fails
            header_text = Paragraph("CSIT Inventory Management System", styles['CustomTitle'])
            elements.append(header_text)
            elements.append(Spacer(1, 10))
        
        # Add title
        title_text = f"{report_type.replace('_', ' ').title()} Report"
        title = Paragraph(title_text, styles['CustomTitle'])
        elements.append(title)
        
        # Add horizontal line
        elements.append(Spacer(1, 0.1*inch))
        elements.append(HRFlowable(
            width="100%",
            thickness=1,
            lineCap='round',
            color=colors.HexColor('#2E5984'),
            spaceAfter=0.3*inch
        ))
        
        # Add report metadata in a two-column layout
        metadata = [
            ["Generated on:", report_data.get('generated_at', 'N/A')],
            ["Total Records:", str(report_data.get('total_records', report_data.get('total_items', 0)))],
        ]
        
        if report_data.get('total_amount'):
            metadata.append(["Total Amount:", f"${report_data['total_amount']:.2f}"])
        if report_data.get('total_value'):
            metadata.append(["Total Value:", f"${report_data['total_value']:.2f}"])
        if report_data.get('total_quantity_discarded'):
            metadata.append(["Total Discarded:", str(report_data['total_quantity_discarded'])])
        
        # Create metadata table
        metadata_table = Table(metadata, colWidths=[2*inch, 3*inch])
        metadata_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1E3F66')),  # Dark blue for labels
            ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        
        elements.append(metadata_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Generate table based on report type
        if report_data.get('data'):
            data = report_data['data']
            if data and len(data) > 0:  # Check if data exists and is not empty
                # Section header
                elements.append(Paragraph("Report Details", styles['SectionHeader']))
                
                # Get headers from first data item
                headers = list(data[0].keys())
                # Convert headers to readable format
                header_labels = [h.replace('_', ' ').title() for h in headers]
                
                # Prepare table data
                table_data = [header_labels]
                for item in data:
                    row = []
                    for header in headers:
                        value = item.get(header, '')
                        # Format dates
                        if isinstance(value, str) and 'date' in header.lower():
                            try:
                                parsed_date = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                value = parsed_date.strftime('%Y-%m-%d')
                            except:
                                pass
                        # Format currency
                        elif isinstance(value, (int, float)) and any(currency_word in header.lower() for currency_word in ['price', 'amount', 'value']):
                            value = f"${value:,.2f}"
                        # Format quantity dict
                        elif header.lower() == 'quantity' and isinstance(value, dict):
                            value = f"Received: {value.get('received', '')}, Issued: {value.get('issued', '')}, Balance: {value.get('balance', '')}"
                        # Format large numbers with commas
                        elif isinstance(value, int) and value > 1000:
                            value = f"{value:,}"
                        # Wrap long text in Paragraph for word wrapping and row height
                        if isinstance(value, str) and (len(value) > 15 or '\n' in value or ',' in value):
                            value = Paragraph(value, styles['Normal'])
                        row.append(value)
                    table_data.append(row)
                
                # Create table with alternating row colors
                page_width = A4[0] - doc.leftMargin - doc.rightMargin
                num_columns = len(table_data[0])
                col_width = page_width / num_columns
                table = Table(table_data, colWidths=[col_width]*num_columns, repeatRows=1)
                table_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5984')),  # Header background
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Default row color
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D3D3D3')),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('WORDWRAP', (0, 0), (-1, -1), 'CJK'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]
                # Dynamically add alternating row colors for even rows (starting from row 2, i.e., index 2)
                for i in range(2, len(table_data), 2):
                    table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F5F5F5')))
                table.setStyle(TableStyle(table_style))
                
                elements.append(table)
            else:
                # Add a message when no data is available
                elements.append(Paragraph("Report Details", styles['SectionHeader']))
                elements.append(Paragraph("No data available for the selected filters.", styles['Metadata']))
        
        # Add footer
        elements.append(Spacer(1, 0.5*inch))
        elements.append(HRFlowable(
            width="100%",
            thickness=0.5,
            lineCap='round',
            color=colors.HexColor('#2E5984'),
            spaceAfter=0.2*inch
        ))
        footer_text = f"Confidential - {datetime.now().year} © NED UET"
        elements.append(Paragraph(footer_text, styles['Footer']))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def _generate_register_pdf_content(self, report_data):
        """Generate specialized PDF content for Register report with proper table layout"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        elements = []
        
        # Custom styles
        styles = getSampleStyleSheet()
        
        # Add custom styles for Register report
        styles.add(ParagraphStyle(
            'RegisterTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=15,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#2E5984')
        ))
        
        styles.add(ParagraphStyle(
            'RegisterHeader',
            parent=styles['Normal'],
            fontSize=9,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
            spaceAfter=2,
            spaceBefore=2,
            leading=11
        ))
        
        styles.add(ParagraphStyle(
            'RegisterSubHeader',
            parent=styles['Normal'],
            fontSize=8,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
            spaceAfter=1,
            spaceBefore=1,
            leading=10
        ))
        
        styles.add(ParagraphStyle(
            'RegisterCell',
            parent=styles['Normal'],
            fontSize=8,
            fontName='Helvetica',
            alignment=TA_CENTER,
            spaceAfter=1,
            spaceBefore=1,
            leading=10
        ))
        
        # Add company logo and header
        try:
            # Try multiple possible paths for the logo
            logo_paths = [
                "nedlogo.jpeg",
                "inventory/views/nedlogo.jpeg",
                os.path.join(os.path.dirname(__file__), "nedlogo.jpeg"),
                os.path.join(settings.BASE_DIR, "inventory", "views", "nedlogo.jpeg")
            ]
            
            logo_found = False
            for logo_path in logo_paths:
                if os.path.exists(logo_path):
                    logo = Image(logo_path, width=1*inch, height=1*inch)
                    elements.append(logo)
                    elements.append(Spacer(1, 10))
                    logo_found = True
                    break
                    
            if not logo_found:
                # Add a text header instead
                header_text = Paragraph("CSIT Inventory Management System", styles['RegisterTitle'])
                elements.append(header_text)
                elements.append(Spacer(1, 10))
        except Exception as e:
            # Add a text header instead if logo fails
            header_text = Paragraph("CSIT Inventory Management System", styles['RegisterTitle'])
            elements.append(header_text)
            elements.append(Spacer(1, 10))
        
        # Add company header
        header_text = Paragraph("NED UNIVERSITY OF ENGINEERING AND TECHNOLOGY, KARACHI", styles['RegisterTitle'])
        elements.append(header_text)
        
        # Add document code
        doc_code = Paragraph("Document Code: F/SOP/SD 01/52/01", styles['RegisterCell'])
        elements.append(doc_code)
        elements.append(Spacer(1, 10))
        
        # Add report metadata
        metadata = []
        register_type = report_data.get('register_type')
        if register_type:
            metadata.append(["Register Type:", f"{register_type} Register"])
        metadata.append(["Generated on:", str(report_data.get('generated_at', 'N/A'))])
        metadata.append(["Total Records:", str(report_data.get('total_records', 0))])
        metadata_table = Table(metadata, colWidths=[1.5*inch, 3*inch])
        metadata_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1E3F66')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        
        elements.append(metadata_table)
        elements.append(Spacer(1, 15))
        
        # Generate Register table
        if report_data.get('data'):
            data = report_data['data']
            if data and len(data) > 0:
                # Define column widths for Register table (11 columns total)
                page_width = A4[0] - doc.leftMargin - doc.rightMargin
                # Calculate total width needed and adjust proportionally
                total_width = 8.8*inch  # Sum of all column widths
                scale_factor = page_width / total_width
                
                col_widths = [
                    0.7*inch * scale_factor,  # Date
                    0.8*inch * scale_factor,  # Received/Issued
                    1.0*inch * scale_factor,  # Voucher No
                    1.1*inch * scale_factor,  # Particulars
                    0.7*inch * scale_factor,  # Unit
                    0.6*inch * scale_factor,  # Unit Price
                    0.8*inch * scale_factor,  # Total Cost
                    0.5*inch * scale_factor,  # Received (Quantity)
                    0.5*inch * scale_factor,  # Issued (Quantity)
                    0.5*inch * scale_factor,  # Balance (Quantity)
                    0.9*inch * scale_factor,  # Remarks
                ]
                
                # Create header rows with proper multi-line headers
                header_row1 = [
                    Paragraph("Date", styles['RegisterHeader']),
                    Paragraph("Received/<br/>Issued", styles['RegisterHeader']),
                    Paragraph("Voucher / Cash Memo /<br/>Requisition / Purchase<br/>Order No.", styles['RegisterHeader']),
                    Paragraph("Particulars", styles['RegisterHeader']),
                    Paragraph("Accounting /<br/>Measuring Unit", styles['RegisterHeader']),
                    Paragraph("Unit Price", styles['RegisterHeader']),
                    Paragraph("Total Cost<br/>(with taxes)", styles['RegisterHeader']),
                    Paragraph("Quantity", styles['RegisterHeader']),
                    Paragraph("", styles['RegisterHeader']),
                    Paragraph("", styles['RegisterHeader']),
                    Paragraph("Remarks / Initials of<br/>Authorized Persons", styles['RegisterHeader']),
                ]
                
                header_row2 = [
                    Paragraph("", styles['RegisterSubHeader']),
                    Paragraph("", styles['RegisterSubHeader']),
                    Paragraph("", styles['RegisterSubHeader']),
                    Paragraph("", styles['RegisterSubHeader']),
                    Paragraph("", styles['RegisterSubHeader']),
                    Paragraph("", styles['RegisterSubHeader']),
                    Paragraph("", styles['RegisterSubHeader']),
                    Paragraph("Received", styles['RegisterSubHeader']),
                    Paragraph("Issued", styles['RegisterSubHeader']),
                    Paragraph("Balance", styles['RegisterSubHeader']),
                    Paragraph("", styles['RegisterSubHeader']),
                ]
                
                # Prepare table data
                table_data = [header_row1, header_row2]
                
                for item in data:
                    row = [
                        Paragraph(item.get('date', ''), styles['RegisterCell']),
                        Paragraph(item.get('receivedIssued', ''), styles['RegisterCell']),
                        Paragraph(item.get('voucherNo', ''), styles['RegisterCell']),
                        Paragraph(item.get('particulars', ''), styles['RegisterCell']),
                        Paragraph(item.get('unit', ''), styles['RegisterCell']),
                        Paragraph(f"${item.get('unitPrice', 0):,.2f}", styles['RegisterCell']),
                        Paragraph(f"${item.get('totalCost', 0):,.2f}", styles['RegisterCell']),
                        Paragraph(str(item.get('quantity', {}).get('received', '')), styles['RegisterCell']),
                        Paragraph(str(item.get('quantity', {}).get('issued', '')), styles['RegisterCell']),
                        Paragraph(str(item.get('quantity', {}).get('balance', '')), styles['RegisterCell']),
                        Paragraph(item.get('remarks', ''), styles['RegisterCell']),
                    ]
                    table_data.append(row)
                
                # Create table
                table = Table(table_data, colWidths=col_widths, repeatRows=2)
                
                # Define table style with proper spacing and formatting
                table_style = [
                    # Header styling (first two rows)
                    ('BACKGROUND', (0, 0), (-1, 1), colors.HexColor('#E0E0E0')),
                    ('TEXTCOLOR', (0, 0), (-1, 1), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 1), 9),
                    ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, 1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, 1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, 1), 6),
                    ('LEFTPADDING', (0, 0), (-1, 1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, 1), 3),
                    
                    # Data row styling
                    ('BACKGROUND', (0, 2), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 2), (-1, -1), colors.black),
                    ('FONTNAME', (0, 2), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 2), (-1, -1), 8),
                    ('ALIGN', (0, 2), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 2), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 2), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 2), (-1, -1), 4),
                    ('LEFTPADDING', (0, 2), (-1, -1), 2),
                    ('RIGHTPADDING', (0, 2), (-1, -1), 2),
                    
                    # Grid and borders
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('BOX', (0, 0), (-1, -1), 1, colors.black),
                    
                    # Word wrapping for all cells
                    ('WORDWRAP', (0, 0), (-1, -1), 'CJK'),
                    
                    # Merge quantity header cells
                    ('SPAN', (7, 0), (9, 0)),  # Merge "Quantity" across 3 columns
                ]
                
                # Add alternating row colors for data rows
                for i in range(3, len(table_data), 2):
                    table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8F9FA')))
                
                table.setStyle(TableStyle(table_style))
                elements.append(table)
            else:
                # Add a message when no data is available
                elements.append(Paragraph("No data available for the selected filters.", styles['RegisterCell']))
        
        # Add footer
        elements.append(Spacer(1, 20))
        footer_text = f"Confidential - {datetime.now().year} © NED UET"
        elements.append(Paragraph(footer_text, styles['RegisterCell']))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def _generate_excel_content(self, report_data, report_type):
        """Generate Excel content based on report type and data"""
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.replace('_', ' ').title()} Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        data_font = Font()
        data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        data_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws['A1'] = f"{report_type.replace('_', ' ').title()} Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Add metadata
        current_row = 3
        ws[f'A{current_row}'] = f"Generated on: {report_data.get('generated_at', 'N/A')}"
        current_row += 1
        ws[f'A{current_row}'] = f"Total Records: {report_data.get('total_records', report_data.get('total_items', 0))}"
        current_row += 1
        
        if report_data.get('total_amount'):
            ws[f'A{current_row}'] = f"Total Amount: ${report_data['total_amount']:.2f}"
            current_row += 1
        if report_data.get('total_value'):
            ws[f'A{current_row}'] = f"Total Value: ${report_data['total_value']:.2f}"
            current_row += 1
        if report_data.get('total_quantity_discarded'):
            ws[f'A{current_row}'] = f"Total Quantity Discarded: {report_data['total_quantity_discarded']}"
            current_row += 1
        
        current_row += 1  # Add space before table
        
        # Generate table based on report type
        if report_data.get('data'):
            data = report_data['data']
            if data and len(data) > 0:  # Check if data exists and is not empty
                # Get headers from first data item
                headers = list(data[0].keys())
                # Convert headers to readable format
                header_labels = [h.replace('_', ' ').title() for h in headers]
                
                # Add headers
                for col, header in enumerate(header_labels, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                current_row += 1
                
                # Add data rows
                for item in data:
                    for col, header in enumerate(headers, 1):
                        value = item.get(header, '')
                        
                        # Format dates
                        if isinstance(value, str) and 'date' in header.lower():
                            try:
                                parsed_date = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                value = parsed_date.strftime('%Y-%m-%d')
                            except:
                                pass
                        # Format currency
                        elif isinstance(value, (int, float)) and any(currency_word in header.lower() for currency_word in ['price', 'amount', 'value']):
                            value = f"${value:.2f}"
                        
                        cell = ws.cell(row=current_row, column=col, value=str(value))
                        cell.font = data_font
                        cell.fill = data_fill
                        cell.alignment = data_alignment
                        cell.border = border
                    
                    current_row += 1
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
            else:
                # Add a message when no data is available
                ws[f'A{current_row}'] = "No data available for the selected filters."
                ws[f'A{current_row}'].font = Font(bold=True, color="FF0000")
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @action(detail=False, methods=['post'])
    def generate_procurement_report(self, request):
        """Generate procurement report with filters"""
        try:
            filters = request.data.get('filters', {})
            
            # Build query based on filters
            queryset = Procurement.objects.all()
            
            start_date = self._parse_date_filter(filters.get('startDate'))
            end_date = self._parse_date_filter(filters.get('endDate'))
            
            if start_date:
                queryset = queryset.filter(created_at__gte=start_date)
            if end_date:
                # For end date, we want to include the entire day
                end_date = timezone.make_aware(
                    datetime.combine(end_date.date(), datetime.max.time()),
                    timezone=timezone.get_current_timezone()
                )
                queryset = queryset.filter(created_at__lte=end_date)
            if filters.get('item'):
                queryset = queryset.filter(items__item__id=filters['item'])
            if filters.get('procurement_type'):
                queryset = queryset.filter(procurement_type=filters['procurement_type'])
            if filters.get('supplier'):
                queryset = queryset.filter(supplier=filters['supplier'])
            
            # Create report record
            report = Report.objects.create(
                report_type='procurement',
                filters=filters,
                generated_by=self._get_generated_by_user(request)
            )
            
            # Prepare report data
            report_data = {
                'report_id': report.id,
                'report_type': 'procurement',
                'generated_at': report.generated_at,
                'filters': filters,
                'total_records': queryset.count(),
                'total_amount': sum(
                    sum(float(item.unit_price) * item.quantity for item in p.items.all())
                    for p in queryset
                ),
                'data': []
            }
            
            for procurement in queryset:
                total_amount = sum(float(item.unit_price) * item.quantity for item in procurement.items.all())
                items_summary = ", ".join(
                    f"{item.item.name} ({item.quantity})" for item in procurement.items.all()
                )
                report_data['data'].append({
                    'order_number': procurement.order_number,
                    'items_summary': items_summary,
                    'supplier': procurement.supplier,
                    'order_date': procurement.order_date,
                    'total_amount': total_amount,
                })
            
            return Response(report_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Failed to generate procurement report: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def generate_stock_movement_report(self, request):
        """Generate stock movement report with filters"""
        try:
            filters = request.data.get('filters', {})
            
            queryset = StockMovement.objects.all()
            
            start_date = self._parse_date_filter(filters.get('startDate'))
            end_date = self._parse_date_filter(filters.get('endDate'))
            
            if start_date:
                queryset = queryset.filter(movement_date__gte=start_date.date())
            if end_date:
                queryset = queryset.filter(movement_date__lte=end_date.date())
            if filters.get('user'):
                queryset = queryset.filter(received_by__id=filters['user'])
            if filters.get('item'):
                queryset = queryset.filter(item__id=filters['item'])
            if filters.get('from_location'):
                queryset = queryset.filter(from_location__id=filters['from_location'])
            if filters.get('to_location'):
                queryset = queryset.filter(to_location__id=filters['to_location'])
            
            report = Report.objects.create(
                report_type='stock_movement',
                filters=filters,
                generated_by=self._get_generated_by_user(request)
            )
            
            report_data = {
                'report_id': report.id,
                'report_type': 'stock_movement',
                'generated_at': report.generated_at,
                'filters': filters,
                'total_records': queryset.count(),
                'data': []
            }
            
            for movement in queryset:
                report_data['data'].append({
                    'item_name': movement.item.name,
                    'quantity': movement.quantity,
                    'from_location': movement.from_location.name,
                    'to_location': movement.to_location.name,
                    'movement_date': movement.movement_date,
                    'received_by': movement.received_by.name,
                    'notes': movement.notes
                })
            
            return Response(report_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Failed to generate stock movement report: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def generate_inventory_report(self, request):
        """Generate inventory report with filters, supporting item, location, and date filters."""
        try:
            filters = request.data.get('filters', {})

            start_date = self._parse_date_filter(filters.get('startDate'))
            end_date = self._parse_date_filter(filters.get('endDate'))

            # If location filter is provided, use InventoryByLocation
            if filters.get('location'):
                inv_qs = InventoryByLocation.objects.all()
                if filters.get('item'):
                    inv_qs = inv_qs.filter(item__id=filters['item'])
                inv_qs = inv_qs.filter(location__id=filters['location'])
                if start_date:
                    inv_qs = inv_qs.filter(last_updated__gte=start_date)
                if end_date:
                    inv_qs = inv_qs.filter(last_updated__lte=end_date)
                report_data = {
                    'report_type': 'inventory',
                    'generated_at': timezone.now(),
                    'filters': filters,
                    'total_items': inv_qs.count(),
                    'total_value': sum(float(inv.item.unit_price) * inv.quantity for inv in inv_qs),
                    'data': []
                }
                for inv in inv_qs:
                    report_data['data'].append({
                        'item_name': inv.item.name,
                        'category': inv.item.category.name,
                        'location': inv.location.name,
                        'quantity': inv.quantity,
                        'unit_price': float(inv.item.unit_price),
                        'total_value': float(inv.item.unit_price) * inv.quantity
                    })
                return Response(report_data, status=status.HTTP_200_OK)

            # If only item filter is provided, show all locations for that item
            elif filters.get('item'):
                inv_qs = InventoryByLocation.objects.filter(item__id=filters['item'])
                if start_date:
                    inv_qs = inv_qs.filter(last_updated__gte=start_date)
                if end_date:
                    inv_qs = inv_qs.filter(last_updated__lte=end_date)
                report_data = {
                    'report_type': 'inventory',
                    'generated_at': timezone.now(),
                    'filters': filters,
                    'total_items': inv_qs.count(),
                    'total_value': sum(float(inv.item.unit_price) * inv.quantity for inv in inv_qs),
                    'data': []
                }
                for inv in inv_qs:
                    report_data['data'].append({
                        'item_name': inv.item.name,
                        'category': inv.item.category.name,
                        'location': inv.location.name,
                        'quantity': inv.quantity,
                        'unit_price': float(inv.item.unit_price),
                        'total_value': float(inv.item.unit_price) * inv.quantity
                    })
                return Response(report_data, status=status.HTTP_200_OK)

            # Default: show all items (old logic)
            queryset = Item.objects.all()
            report = Report.objects.create(
                report_type='inventory',
                filters=filters,
                generated_by=self._get_generated_by_user(request)
            )
            report_data = {
                'report_id': report.id,
                'report_type': 'inventory',
                'generated_at': report.generated_at,
                'filters': filters,
                'total_items': queryset.count(),
                'total_value': sum(float(item.unit_price) * sum(inv.quantity for inv in item.inventory_by_location.all()) for item in queryset),
                'data': []
            }
            for item in queryset:
                total_quantity = sum(inv.quantity for inv in item.inventory_by_location.all())
                locations = ', '.join(inv.location.name for inv in item.inventory_by_location.all())
                report_data['data'].append({
                    'item_name': item.name,
                    'category': item.category.name,
                    'location': locations,
                    'quantity': total_quantity,
                    'unit_price': float(item.unit_price),
                    'total_value': float(item.unit_price) * total_quantity
                })
            return Response(report_data, status=status.HTTP_200_OK)
        except Exception as e:
            import traceback
            print("INVENTORY REPORT ERROR:", e)
            traceback.print_exc()
            return Response(
                {'error': f'Failed to generate inventory report: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def generate_stock_requests_report(self, request):
        """Generate stock requests report with filters"""
        try:
            filters = request.data.get('filters', {})
            
            queryset = SendingStockRequest.objects.all()
            
            start_date = self._parse_date_filter(filters.get('startDate'))
            end_date = self._parse_date_filter(filters.get('endDate'))
            
            if start_date:
                queryset = queryset.filter(created_at__gte=start_date)
            if end_date:
                # For end date, we want to include the entire day
                end_date = timezone.make_aware(
                    datetime.combine(end_date.date(), datetime.max.time()),
                    timezone=timezone.get_current_timezone()
                )
                queryset = queryset.filter(created_at__lte=end_date)
            if filters.get('status'):
                queryset = queryset.filter(status=filters['status'])
            if filters.get('user'):
                queryset = queryset.filter(requested_by__id=filters['user'])
            if filters.get('item'):
                queryset = queryset.filter(item__id=filters['item'])
            
            report = Report.objects.create(
                report_type='stock_requests',
                filters=filters,
                generated_by=self._get_generated_by_user(request)
            )
            
            report_data = {
                'report_id': report.id,
                'report_type': 'stock_requests',
                'generated_at': report.generated_at,
                'filters': filters,
                'total_records': queryset.count(),
                'pending_requests': queryset.filter(status='Pending').count(),
                'approved_requests': queryset.filter(status='Approved').count(),
                'rejected_requests': queryset.filter(status='Rejected').count(),
                'data': []
            }
            
            for request_item in queryset:
                report_data['data'].append({
                    'item_name': request_item.item.name,
                    'quantity': request_item.quantity,
                    'status': request_item.status,
                    'requested_by': str(request_item.requested_by) if request_item.requested_by else 'Unknown',
                    'created_at': request_item.created_at
                })
            
            return Response(report_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Failed to generate stock requests report: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def generate_discarded_items_report(self, request):
        """Generate discarded items report with filters"""
        try:
            filters = request.data.get('filters', {})
            
            queryset = DiscardedItem.objects.all()
            
            start_date = self._parse_date_filter(filters.get('startDate'))
            end_date = self._parse_date_filter(filters.get('endDate'))
            
            if start_date:
                queryset = queryset.filter(date__gte=start_date.date())
            if end_date:
                queryset = queryset.filter(date__lte=end_date.date())
            if filters.get('reason'):
                queryset = queryset.filter(reason=filters['reason'])
            if filters.get('user'):
                queryset = queryset.filter(discarded_by__id=filters['user'])
            if filters.get('item'):
                queryset = queryset.filter(item__id=filters['item'])
            if filters.get('location'):
                queryset = queryset.filter(location__id=filters['location'])
            
            report = Report.objects.create(
                report_type='discarded_items',
                filters=filters,
                generated_by=self._get_generated_by_user(request)
            )
            
            report_data = {
                'report_id': report.id,
                'report_type': 'discarded_items',
                'generated_at': report.generated_at,
                'filters': filters,
                'total_records': queryset.count(),
                'total_quantity_discarded': sum(item.quantity for item in queryset),
                'data': []
            }
            
            for discarded_item in queryset:
                report_data['data'].append({
                    'id': discarded_item.id,
                    'item_name': discarded_item.item.name,
                    'location': discarded_item.location.name,
                    'quantity': discarded_item.quantity,
                    'date': discarded_item.date,
                    'reason': discarded_item.reason,
                    'discarded_by': discarded_item.discarded_by.name if discarded_item.discarded_by else 'Unknown',
                })
            
            return Response(report_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Failed to generate discarded items report: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def export_pdf(self, request, pk=None):
        """Export report as PDF"""
        try:
            # Get the report object
            try:
                report = self.get_object()
            except Exception as e:
                print(f"Error getting report object: {str(e)}")
                # Try to get the report directly using pk
                if pk:
                    try:
                        report = Report.objects.get(id=pk)
                    except Report.DoesNotExist:
                        return Response(
                            {'error': f'Report with ID {pk} not found'}, 
                            status=status.HTTP_404_NOT_FOUND
                        )
                else:
                    return Response(
                        {'error': 'Report ID not provided'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
            # Get the report data based on report type
            report_data = None
            filters = report.filters or {}
            print(f"Exporting PDF for report {report.id} of type {report.report_type}")
            print(f"Filters: {filters}")

            if report.report_type == 'register':
                # Import time at the top level
                from datetime import time
                
                # Get main store location
                from ..models import Location, ProcurementItem
                try:
                    main_store = Location.get_main_store()
                    if not main_store:
                        return Response(
                            {'error': 'Main store location not found'}, 
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR
                        )
                except Exception as e:
                    return Response(
                        {'error': f'Failed to get main store location: {str(e)}'}, 
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
                
                # Collect all data entries (both stock movements and procurements)
                all_entries = []
                
                # Parse date filters once at the beginning
                start_date = self._parse_date_filter(filters.get('startDate'))
                end_date = self._parse_date_filter(filters.get('endDate'))
                
                # 1. Get StockMovement records for items in the selected category
                try:
                    stock_movements = StockMovement.objects.filter(
                        item__category__id=filters.get('category')
                    ).order_by('movement_date', 'id')
                    
                    # Apply date filters to stock movements
                    if start_date:
                        stock_movements = stock_movements.filter(movement_date__gte=start_date.date())
                    if end_date:
                        stock_movements = stock_movements.filter(movement_date__lte=end_date.date())
                    
                    # Filter for movements involving main store (either from or to)
                    stock_movements = stock_movements.filter(
                        Q(from_location=main_store) | Q(to_location=main_store)
                    )
                    
                    # Convert stock movements to register entries
                    for movement in stock_movements:
                        # Determine if this is a receipt or issue
                        if movement.to_location == main_store:
                            # Receipt into main store
                            received_issued = 'Received'
                            received = movement.quantity
                            issued = ''
                            balance_change = movement.quantity  # Positive for received
                            # Create descriptive remarks for received items
                            remarks = movement.notes or f"Stock movement from {movement.from_location.name} to {movement.to_location.name}"
                        elif movement.from_location == main_store:
                            # Issue from main store
                            received_issued = 'Issued'
                            received = ''
                            issued = movement.quantity
                            balance_change = -movement.quantity  # Negative for issued
                            # Create descriptive remarks for issued items
                            remarks = movement.notes or f"Stock movement from {movement.from_location.name} to {movement.to_location.name}"
                        else:
                            continue
                        
                        # Get voucher number from movement notes or create one
                        voucher_no = movement.notes or f"MOV-{movement.id:06d}"
                        
                        # Create datetime for proper ordering - use morning time for received, afternoon for issued
                        if received_issued == 'Received':
                            # Received entries get morning time (9 AM)
                            movement_time = time(hour=9, minute=0, second=0)
                        else:
                            # Issued entries get afternoon time (2 PM) to ensure they come after received entries
                            movement_time = time(hour=14, minute=0, second=0)
                        
                        movement_datetime = timezone.make_aware(
                            datetime.combine(movement.movement_date, movement_time),
                            timezone=timezone.get_current_timezone()
                        )
                        
                        all_entries.append({
                            'datetime': movement_datetime,  # For sorting
                            'date': movement.movement_date.strftime('%Y-%m-%d'),
                            'receivedIssued': received_issued,
                            'voucherNo': voucher_no,
                            'particulars': movement.item.name,
                            'unit': getattr(movement.item, 'unit', 'PCS'),
                            'unitPrice': float(movement.item.unit_price),
                            'totalCost': float(movement.item.unit_price) * movement.quantity,
                            'quantity': {
                                'received': received if received else '',
                                'issued': issued if issued else '',
                                'balance': 0,  # Will be calculated later
                            },
                            'remarks': remarks,
                            'item_id': movement.item_id,
                            'balance_change': balance_change,  # For balance calculation
                        })
                except Exception as e:
                    print(f"Error processing stock movements in PDF export: {str(e)}")
                    # Continue with empty stock movements if there's an error
                    pass
                
                # 2. Get Procurement records for items in the selected category
                try:
                    procurements = Procurement.objects.filter(
                        items__item__category__id=filters.get('category')
                    ).order_by('order_date', 'id')
                    
                    # Apply date filters to procurements (use order_date for consistency)
                    if start_date:
                        procurements = procurements.filter(order_date__gte=start_date.date())
                    if end_date:
                        procurements = procurements.filter(order_date__lte=end_date.date())
                    
                    # Convert procurements to register entries
                    for procurement in procurements:
                        for proc_item in procurement.items.all():
                            if proc_item.item.category.id == int(filters.get('category')):
                                # This is a receipt into main store from procurement
                                voucher_no = f"PO-{procurement.order_number}"
                                
                                # Use morning time (8 AM) for procurement entries to ensure they come first
                                procurement_time = timezone.make_aware(
                                    datetime.combine(procurement.order_date, time(hour=8, minute=0, second=0)),
                                    timezone=timezone.get_current_timezone()
                                )
                                
                                all_entries.append({
                                    'datetime': procurement_time,  # For chronological ordering
                                    'date': procurement.order_date.strftime('%Y-%m-%d'),
                                    'receivedIssued': 'Received',
                                    'voucherNo': voucher_no,
                                    'particulars': proc_item.item.name,
                                    'unit': getattr(proc_item.item, 'unit', 'PCS'),
                                    'unitPrice': float(proc_item.unit_price),
                                    'totalCost': float(proc_item.unit_price) * proc_item.quantity,
                                    'quantity': {
                                        'received': proc_item.quantity,
                                        'issued': '',
                                        'balance': 0,  # Will be calculated later
                                    },
                                    'remarks': f"Procurement from {procurement.supplier}",
                                    'item_id': proc_item.item_id,
                                    'balance_change': proc_item.quantity,  # Positive for received
                                })
                except Exception as e:
                    print(f"Error processing procurements in PDF export: {str(e)}")
                    # Continue with empty procurements if there's an error
                    pass
                
                # Check if we have any data
                if not all_entries:
                    return Response(
                        {'error': 'No data found for the selected category and date range'}, 
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Sort all entries by exact datetime for proper chronological order
                all_entries.sort(key=lambda x: (x['datetime'], x['receivedIssued'] == 'Issued'))
                
                # Calculate running balance per item
                balance_map = {}  # {item_id: balance}
                data = []
                
                for entry in all_entries:
                    try:
                        item_id = entry['item_id']
                        prev_balance = balance_map.get(item_id, 0)
                        
                        # Calculate the new balance
                        new_balance = prev_balance + entry['balance_change']
                        balance_map[item_id] = new_balance
                        
                        # Update the balance in the entry
                        entry['quantity']['balance'] = new_balance
                        
                        # Remove helper fields
                        del entry['datetime']
                        del entry['item_id']
                        del entry['balance_change']
                        
                        data.append(entry)
                    except Exception as e:
                        print(f"Error processing entry in PDF export: {str(e)}")
                        # Skip this entry if there's an error
                        continue
                
                # Fetch category name for Register Type
                category_name = None
                category_id = filters.get('category')
                if category_id:
                    try:
                        category_obj = Category.objects.get(id=category_id)
                        category_name = category_obj.name
                    except Category.DoesNotExist:
                        category_name = "Unknown"
                else:
                    category_name = "Unknown"
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'register',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_records': len(data),
                    'data': data,
                    'register_type': category_name,  # Add register type for PDF
                }
                pdf_buffer = self._generate_pdf_content(report_data, 'register')
                filename = f"register_report_{report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            elif report.report_type == 'procurement':
                queryset = Procurement.objects.all()
                if filters.get('startDate'):
                    start_date = self._parse_date_filter(filters['startDate'])
                    if start_date:
                        queryset = queryset.filter(created_at__gte=start_date)
                if filters.get('endDate'):
                    end_date = self._parse_date_filter(filters['endDate'])
                    if end_date:
                        end_date = timezone.make_aware(
                            datetime.combine(end_date.date(), datetime.max.time()),
                            timezone=timezone.get_current_timezone()
                        )
                        queryset = queryset.filter(created_at__lte=end_date)
                if filters.get('item'):
                    queryset = queryset.filter(items__item__id=filters['item'])
                
                print(f"Procurement queryset count: {queryset.count()}")
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'procurement',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_records': queryset.count(),
                    'total_amount': sum(
                        sum(float(item.unit_price) * item.quantity for item in p.items.all())
                        for p in queryset
                    ),
                    'data': []
                }
                
                for procurement in queryset:
                    total_amount = sum(float(item.unit_price) * item.quantity for item in procurement.items.all())
                    items_summary = ", ".join(
                        f"{item.item.name} ({item.quantity})" for item in procurement.items.all()
                    )
                    report_data['data'].append({
                        'order_number': procurement.order_number,
                        'items_summary': items_summary,
                        'supplier': procurement.supplier,
                        'order_date': procurement.order_date,
                        'total_amount': total_amount,
                    })
            
            elif report.report_type == 'stock_movement':
                queryset = StockMovement.objects.all()
                if filters.get('startDate'):
                    start_date = self._parse_date_filter(filters['startDate'])
                    if start_date:
                        queryset = queryset.filter(movement_date__gte=start_date.date())
                if filters.get('endDate'):
                    end_date = self._parse_date_filter(filters['endDate'])
                    if end_date:
                        queryset = queryset.filter(movement_date__lte=end_date.date())
                if filters.get('user'):
                    queryset = queryset.filter(received_by__id=filters['user'])
                if filters.get('item'):
                    queryset = queryset.filter(item__id=filters['item'])
                if filters.get('from_location'):
                    queryset = queryset.filter(from_location__id=filters['from_location'])
                if filters.get('to_location'):
                    queryset = queryset.filter(to_location__id=filters['to_location'])
                
                print(f"Stock movement queryset count: {queryset.count()}")
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'stock_movement',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_records': queryset.count(),
                    'data': []
                }
                
                for movement in queryset:
                    report_data['data'].append({
                        'item_name': movement.item.name,
                        'quantity': movement.quantity,
                        'from_location': movement.from_location.name,
                        'to_location': movement.to_location.name,
                        'movement_date': movement.movement_date,
                        'received_by': movement.received_by.name,
                        'notes': movement.notes
                    })
            
            elif report.report_type == 'inventory':
                queryset = Item.objects.all()
                if filters.get('item'):
                    queryset = queryset.filter(id=filters['item'])
                
                print(f"Inventory queryset count: {queryset.count()}")
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'inventory',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_items': queryset.count(),
                    'total_value': sum(float(item.unit_price) * sum(inv.quantity for inv in item.inventory_by_location.all()) for item in queryset),
                    'data': []
                }
                
                for item in queryset:
                    total_quantity = sum(inv.quantity for inv in item.inventory_by_location.all())
                    locations = ', '.join(inv.location.name for inv in item.inventory_by_location.all())
                    report_data['data'].append({
                        'item_name': item.name,
                        'category': item.category.name,
                        'location': locations,
                        'quantity': total_quantity,
                        'unit_price': float(item.unit_price),
                        'total_value': float(item.unit_price) * total_quantity
                    })
            
            elif report.report_type == 'stock_requests':
                queryset = SendingStockRequest.objects.all()
                if filters.get('startDate'):
                    start_date = self._parse_date_filter(filters['startDate'])
                    if start_date:
                        queryset = queryset.filter(created_at__gte=start_date)
                if filters.get('endDate'):
                    end_date = self._parse_date_filter(filters['endDate'])
                    if end_date:
                        end_date = timezone.make_aware(
                            datetime.combine(end_date.date(), datetime.max.time()),
                            timezone=timezone.get_current_timezone()
                        )
                        queryset = queryset.filter(created_at__lte=end_date)
                if filters.get('status'):
                    queryset = queryset.filter(status=filters['status'])
                if filters.get('user'):
                    queryset = queryset.filter(requested_by__id=filters['user'])
                if filters.get('item'):
                    queryset = queryset.filter(item__id=filters['item'])
                
                print(f"Stock requests queryset count: {queryset.count()}")
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'stock_requests',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_records': queryset.count(),
                    'pending_requests': queryset.filter(status='Pending').count(),
                    'approved_requests': queryset.filter(status='Approved').count(),
                    'rejected_requests': queryset.filter(status='Rejected').count(),
                    'data': []
                }
                
                for request_item in queryset:
                    report_data['data'].append({
                        'item_name': request_item.item.name,
                        'quantity': request_item.quantity,
                        'status': request_item.status,
                        'requested_by': str(request_item.requested_by) if request_item.requested_by else 'Unknown',
                        'created_at': request_item.created_at
                    })
            
            elif report.report_type == 'discarded_items':
                queryset = DiscardedItem.objects.all()
                if filters.get('startDate'):
                    start_date = self._parse_date_filter(filters['startDate'])
                    if start_date:
                        queryset = queryset.filter(date__gte=start_date.date())
                if filters.get('endDate'):
                    end_date = self._parse_date_filter(filters['endDate'])
                    if end_date:
                        queryset = queryset.filter(date__lte=end_date.date())
                if filters.get('reason'):
                    queryset = queryset.filter(reason=filters['reason'])
                if filters.get('user'):
                    queryset = queryset.filter(discarded_by__id=filters['user'])
                if filters.get('item'):
                    queryset = queryset.filter(item__id=filters['item'])
                if filters.get('location'):
                    queryset = queryset.filter(location__id=filters['location'])
                
                print(f"Discarded items queryset count: {queryset.count()}")
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'discarded_items',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_records': queryset.count(),
                    'total_quantity_discarded': sum(item.quantity for item in queryset),
                    'data': []
                }
                
                for discarded_item in queryset:
                    report_data['data'].append({
                        'id': discarded_item.id,
                        'item_name': discarded_item.item.name,
                        'location': discarded_item.location.name,
                        'quantity': discarded_item.quantity,
                        'date': discarded_item.date,
                        'reason': discarded_item.reason,
                        'discarded_by': discarded_item.discarded_by.name if discarded_item.discarded_by else 'Unknown',
                    })
            
            elif report.report_type == 'categories':
                category_id = filters.get('category')
                if not category_id:
                    return Response({'error': 'Category filter is required.'}, status=status.HTTP_400_BAD_REQUEST)
                
                items = Item.objects.filter(category__id=category_id)
                
                print(f"Categories report queryset count: {items.count()}")
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'categories',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_items': items.count(),
                    'data': []
                }
                
                for item in items:
                    total_quantity = sum(inv.quantity for inv in item.inventory_by_location.all())
                    locations = ', '.join(inv.location.name for inv in item.inventory_by_location.all())
                    report_data['data'].append({
                        'id': item.id,
                        'category_name': item.category.name,
                        'item_name': item.name,
                        'item_count': total_quantity,
                        'unit_price': float(item.unit_price),
                        'total_value': float(item.unit_price) * total_quantity,
                        'locations': locations,
                    })
            
            print(f"Report data prepared: {len(report_data.get('data', [])) if report_data else 0} items")
            
            if not report_data:
                return Response(
                    {'error': 'Report data not found'}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Generate PDF
            pdf_buffer = self._generate_pdf_content(report_data, report.report_type)
            
            # Create filename
            filename = f"{report.report_type}_report_{report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            # Create HTTP response with PDF
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            print(f"Error in export_pdf: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Failed to export PDF: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def export_excel(self, request, pk=None):
        """Export report as Excel"""
        try:
            # Get the report object
            try:
                report = self.get_object()
            except Exception as e:
                print(f"Error getting report object: {str(e)}")
                # Try to get the report directly using pk
                if pk:
                    try:
                        report = Report.objects.get(id=pk)
                    except Report.DoesNotExist:
                        return Response(
                            {'error': f'Report with ID {pk} not found'}, 
                            status=status.HTTP_404_NOT_FOUND
                        )
                else:
                    return Response(
                        {'error': 'Report ID not provided'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
            # Get the report data based on report type
            report_data = None
            filters = report.filters or {}
            print(f"Exporting Excel for report {report.id} of type {report.report_type}")
            print(f"Filters: {filters}")

            if report.report_type == 'register':
                # Import time at the top level
                from datetime import time
                
                # Get main store location
                from ..models import Location, ProcurementItem
                try:
                    main_store = Location.get_main_store()
                    if not main_store:
                        return Response(
                            {'error': 'Main store location not found'}, 
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR
                        )
                except Exception as e:
                    return Response(
                        {'error': f'Failed to get main store location: {str(e)}'}, 
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
                
                # Collect all data entries (both stock movements and procurements)
                all_entries = []
                
                # Parse date filters once at the beginning
                start_date = self._parse_date_filter(filters.get('startDate'))
                end_date = self._parse_date_filter(filters.get('endDate'))
                
                # 1. Get StockMovement records for items in the selected category
                try:
                    stock_movements = StockMovement.objects.filter(
                        item__category__id=filters.get('category')
                    ).order_by('movement_date', 'id')
                    
                    # Apply date filters to stock movements
                    if start_date:
                        stock_movements = stock_movements.filter(movement_date__gte=start_date.date())
                    if end_date:
                        stock_movements = stock_movements.filter(movement_date__lte=end_date.date())
                    
                    # Filter for movements involving main store (either from or to)
                    stock_movements = stock_movements.filter(
                        Q(from_location=main_store) | Q(to_location=main_store)
                    )
                    
                    # Convert stock movements to register entries
                    for movement in stock_movements:
                        # Determine if this is a receipt or issue
                        if movement.to_location == main_store:
                            # Receipt into main store
                            received_issued = 'Received'
                            received = movement.quantity
                            issued = ''
                            balance_change = movement.quantity  # Positive for received
                            # Create descriptive remarks for received items
                            remarks = movement.notes or f"Stock movement from {movement.from_location.name} to {movement.to_location.name}"
                        elif movement.from_location == main_store:
                            # Issue from main store
                            received_issued = 'Issued'
                            received = ''
                            issued = movement.quantity
                            balance_change = -movement.quantity  # Negative for issued
                            # Create descriptive remarks for issued items
                            remarks = movement.notes or f"Stock movement from {movement.from_location.name} to {movement.to_location.name}"
                        else:
                            continue
                        
                        # Get voucher number from movement notes or create one
                        voucher_no = movement.notes or f"MOV-{movement.id:06d}"
                        
                        # Create datetime for proper ordering - use morning time for received, afternoon for issued
                        if received_issued == 'Received':
                            # Received entries get morning time (9 AM)
                            movement_time = time(hour=9, minute=0, second=0)
                        else:
                            # Issued entries get afternoon time (2 PM) to ensure they come after received entries
                            movement_time = time(hour=14, minute=0, second=0)
                        
                        movement_datetime = timezone.make_aware(
                            datetime.combine(movement.movement_date, movement_time),
                            timezone=timezone.get_current_timezone()
                        )
                        
                        all_entries.append({
                            'datetime': movement_datetime,  # For sorting
                            'date': movement.movement_date.strftime('%Y-%m-%d'),
                            'receivedIssued': received_issued,
                            'voucherNo': voucher_no,
                            'particulars': movement.item.name,
                            'unit': getattr(movement.item, 'unit', 'PCS'),
                            'unitPrice': float(movement.item.unit_price),
                            'totalCost': float(movement.item.unit_price) * movement.quantity,
                            'quantity': {
                                'received': received if received else '',
                                'issued': issued if issued else '',
                                'balance': 0,  # Will be calculated later
                            },
                            'remarks': remarks,
                            'item_id': movement.item_id,
                            'balance_change': balance_change,  # For balance calculation
                        })
                except Exception as e:
                    print(f"Error processing stock movements in Excel export: {str(e)}")
                    # Continue with empty stock movements if there's an error
                    pass
                
                # 2. Get Procurement records for items in the selected category
                try:
                    procurements = Procurement.objects.filter(
                        items__item__category__id=filters.get('category')
                    ).order_by('order_date', 'id')
                    
                    # Apply date filters to procurements (use order_date for consistency)
                    if start_date:
                        procurements = procurements.filter(order_date__gte=start_date.date())
                    if end_date:
                        procurements = procurements.filter(order_date__lte=end_date.date())
                    
                    # Convert procurements to register entries
                    for procurement in procurements:
                        for proc_item in procurement.items.all():
                            if proc_item.item.category.id == int(filters.get('category')):
                                # This is a receipt into main store from procurement
                                voucher_no = f"PO-{procurement.order_number}"
                                
                                # Use morning time (8 AM) for procurement entries to ensure they come first
                                procurement_time = timezone.make_aware(
                                    datetime.combine(procurement.order_date, time(hour=8, minute=0, second=0)),
                                    timezone=timezone.get_current_timezone()
                                )
                                
                                all_entries.append({
                                    'datetime': procurement_time,  # For chronological ordering
                                    'date': procurement.order_date.strftime('%Y-%m-%d'),
                                    'receivedIssued': 'Received',
                                    'voucherNo': voucher_no,
                                    'particulars': proc_item.item.name,
                                    'unit': getattr(proc_item.item, 'unit', 'PCS'),
                                    'unitPrice': float(proc_item.unit_price),
                                    'totalCost': float(proc_item.unit_price) * proc_item.quantity,
                                    'quantity': {
                                        'received': proc_item.quantity,
                                        'issued': '',
                                        'balance': 0,  # Will be calculated later
                                    },
                                    'remarks': f"Procurement from {procurement.supplier}",
                                    'item_id': proc_item.item_id,
                                    'balance_change': proc_item.quantity,  # Positive for received
                                })
                except Exception as e:
                    print(f"Error processing procurements in Excel export: {str(e)}")
                    # Continue with empty procurements if there's an error
                    pass
                
                # Check if we have any data
                if not all_entries:
                    return Response(
                        {'error': 'No data found for the selected category and date range'}, 
                        status=status.HTTP_404_NOT_FOUND
                    )
                
                # Sort all entries by exact datetime for proper chronological order
                all_entries.sort(key=lambda x: (x['datetime'], x['receivedIssued'] == 'Issued'))
                
                # Debug: Print the order of entries
                print("DEBUG: Sorted entries order:")
                for i, entry in enumerate(all_entries):
                    print(f"  {i+1}. {entry['date']} - {entry['receivedIssued']} - {entry['particulars']} - Qty: {entry['balance_change']}")
                
                # Calculate running balance per item
                balance_map = {}  # {item_id: balance}
                data = []
                
                for entry in all_entries:
                    try:
                        item_id = entry['item_id']
                        prev_balance = balance_map.get(item_id, 0)
                        
                        # Calculate the new balance
                        new_balance = prev_balance + entry['balance_change']
                        balance_map[item_id] = new_balance
                        
                        # Update the balance in the entry
                        entry['quantity']['balance'] = new_balance
                        
                        # Remove helper fields
                        del entry['datetime']
                        del entry['item_id']
                        del entry['balance_change']
                        
                        data.append(entry)
                    except Exception as e:
                        print(f"Error processing entry in Excel export: {str(e)}")
                        # Skip this entry if there's an error
                        continue
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'register',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_records': len(data),
                    'data': data
                }
                excel_buffer = self._generate_excel_content(report_data, 'register')
                filename = f"register_report_{report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response = HttpResponse(excel_buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            elif report.report_type == 'procurement':
                queryset = Procurement.objects.all()
                if filters.get('startDate'):
                    start_date = self._parse_date_filter(filters['startDate'])
                    if start_date:
                        queryset = queryset.filter(created_at__gte=start_date)
                if filters.get('endDate'):
                    end_date = self._parse_date_filter(filters['endDate'])
                    if end_date:
                        end_date = timezone.make_aware(
                            datetime.combine(end_date.date(), datetime.max.time()),
                            timezone=timezone.get_current_timezone()
                        )
                        queryset = queryset.filter(created_at__lte=end_date)
                if filters.get('item'):
                    queryset = queryset.filter(items__item__id=filters['item'])
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'procurement',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_records': queryset.count(),
                    'total_amount': sum(
                        sum(float(item.unit_price) * item.quantity for item in p.items.all())
                        for p in queryset
                    ),
                    'data': []
                }
                
                for procurement in queryset:
                    total_amount = sum(float(item.unit_price) * item.quantity for item in procurement.items.all())
                    items_summary = ", ".join(
                        f"{item.item.name} ({item.quantity})" for item in procurement.items.all()
                    )
                    report_data['data'].append({
                        'order_number': procurement.order_number,
                        'items_summary': items_summary,
                        'supplier': procurement.supplier,
                        'order_date': procurement.order_date,
                        'total_amount': total_amount,
                    })
            
            elif report.report_type == 'stock_movement':
                queryset = StockMovement.objects.all()
                if filters.get('startDate'):
                    start_date = self._parse_date_filter(filters['startDate'])
                    if start_date:
                        queryset = queryset.filter(movement_date__gte=start_date.date())
                if filters.get('endDate'):
                    end_date = self._parse_date_filter(filters['endDate'])
                    if end_date:
                        queryset = queryset.filter(movement_date__lte=end_date.date())
                if filters.get('user'):
                    queryset = queryset.filter(received_by__id=filters['user'])
                if filters.get('item'):
                    queryset = queryset.filter(item__id=filters['item'])
                if filters.get('from_location'):
                    queryset = queryset.filter(from_location__id=filters['from_location'])
                if filters.get('to_location'):
                    queryset = queryset.filter(to_location__id=filters['to_location'])
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'stock_movement',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_records': queryset.count(),
                    'data': []
                }
                
                for movement in queryset:
                    report_data['data'].append({
                        'item_name': movement.item.name,
                        'quantity': movement.quantity,
                        'from_location': movement.from_location.name,
                        'to_location': movement.to_location.name,
                        'movement_date': movement.movement_date,
                        'received_by': movement.received_by.name,
                        'notes': movement.notes
                    })
            
            elif report.report_type == 'inventory':
                queryset = Item.objects.all()
                if filters.get('item'):
                    queryset = queryset.filter(id=filters['item'])
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'inventory',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_items': queryset.count(),
                    'total_value': sum(float(item.unit_price) * sum(inv.quantity for inv in item.inventory_by_location.all()) for item in queryset),
                    'data': []
                }
                
                for item in queryset:
                    total_quantity = sum(inv.quantity for inv in item.inventory_by_location.all())
                    locations = ', '.join(inv.location.name for inv in item.inventory_by_location.all())
                    report_data['data'].append({
                        'item_name': item.name,
                        'category': item.category.name,
                        'location': locations,
                        'quantity': total_quantity,
                        'unit_price': float(item.unit_price),
                        'total_value': float(item.unit_price) * total_quantity
                    })
            
            elif report.report_type == 'stock_requests':
                queryset = SendingStockRequest.objects.all()
                if filters.get('startDate'):
                    start_date = self._parse_date_filter(filters['startDate'])
                    if start_date:
                        queryset = queryset.filter(created_at__gte=start_date)
                if filters.get('endDate'):
                    end_date = self._parse_date_filter(filters['endDate'])
                    if end_date:
                        end_date = timezone.make_aware(
                            datetime.combine(end_date.date(), datetime.max.time()),
                            timezone=timezone.get_current_timezone()
                        )
                        queryset = queryset.filter(created_at__lte=end_date)
                if filters.get('status'):
                    queryset = queryset.filter(status=filters['status'])
                if filters.get('user'):
                    queryset = queryset.filter(requested_by__id=filters['user'])
                if filters.get('item'):
                    queryset = queryset.filter(item__id=filters['item'])
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'stock_requests',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_records': queryset.count(),
                    'pending_requests': queryset.filter(status='Pending').count(),
                    'approved_requests': queryset.filter(status='Approved').count(),
                    'rejected_requests': queryset.filter(status='Rejected').count(),
                    'data': []
                }
                
                for request_item in queryset:
                    report_data['data'].append({
                        'item_name': request_item.item.name,
                        'quantity': request_item.quantity,
                        'status': request_item.status,
                        'requested_by': str(request_item.requested_by) if request_item.requested_by else 'Unknown',
                        'created_at': request_item.created_at
                    })
            
            elif report.report_type == 'discarded_items':
                queryset = DiscardedItem.objects.all()
                if filters.get('startDate'):
                    start_date = self._parse_date_filter(filters['startDate'])
                    if start_date:
                        queryset = queryset.filter(date__gte=start_date.date())
                if filters.get('endDate'):
                    end_date = self._parse_date_filter(filters['endDate'])
                    if end_date:
                        queryset = queryset.filter(date__lte=end_date.date())
                if filters.get('reason'):
                    queryset = queryset.filter(reason=filters['reason'])
                if filters.get('user'):
                    queryset = queryset.filter(discarded_by__id=filters['user'])
                if filters.get('item'):
                    queryset = queryset.filter(item__id=filters['item'])
                if filters.get('location'):
                    queryset = queryset.filter(location__id=filters['location'])
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'discarded_items',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_records': queryset.count(),
                    'total_quantity_discarded': sum(item.quantity for item in queryset),
                    'data': []
                }
                
                for discarded_item in queryset:
                    report_data['data'].append({
                        'id': discarded_item.id,
                        'item_name': discarded_item.item.name,
                        'location': discarded_item.location.name,
                        'quantity': discarded_item.quantity,
                        'date': discarded_item.date,
                        'reason': discarded_item.reason,
                        'discarded_by': discarded_item.discarded_by.name if discarded_item.discarded_by else 'Unknown',
                    })
            
            elif report.report_type == 'categories':
                category_id = filters.get('category')
                if not category_id:
                    return Response({'error': 'Category filter is required.'}, status=status.HTTP_400_BAD_REQUEST)
                
                items = Item.objects.filter(category__id=category_id)
                
                report_data = {
                    'report_id': report.id,
                    'report_type': 'categories',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_items': items.count(),
                    'data': []
                }
                
                for item in items:
                    total_quantity = sum(inv.quantity for inv in item.inventory_by_location.all())
                    locations = ', '.join(inv.location.name for inv in item.inventory_by_location.all())
                    report_data['data'].append({
                        'id': item.id,
                        'category_name': item.category.name,
                        'item_name': item.name,
                        'item_count': total_quantity,
                        'unit_price': float(item.unit_price),
                        'total_value': float(item.unit_price) * total_quantity,
                        'locations': locations,
                    })
            
            if not report_data:
                return Response(
                    {'error': 'Report data not found'}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Generate Excel
            excel_buffer = self._generate_excel_content(report_data, report.report_type)
            
            # Create filename
            filename = f"{report.report_type}_report_{report.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Create HTTP response with Excel
            response = HttpResponse(excel_buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Failed to export Excel: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def generate_categories_report(self, request):
        """Generate a report of items in a selected category, with inventory details."""
        try:
            filters = request.data.get('filters', {})
            category_id = filters.get('category')
            if not category_id:
                return Response({'error': 'Category filter is required.'}, status=status.HTTP_400_BAD_REQUEST)

            items = Item.objects.filter(category__id=category_id)
            
            # Create report record
            report = Report.objects.create(
                report_type='categories',
                filters=filters,
                generated_by=self._get_generated_by_user(request)
            )
            
            report_data = {
                'report_id': report.id,
                'report_type': 'categories',
                'generated_at': report.generated_at,
                'filters': filters,
                'total_items': items.count(),
                'data': []
            }
            for item in items:
                total_quantity = sum(inv.quantity for inv in item.inventory_by_location.all())
                locations = ', '.join(inv.location.name for inv in item.inventory_by_location.all())
                report_data['data'].append({
                    'id': item.id,
                    'category_name': item.category.name,
                    'item_name': item.name,
                    'item_count': total_quantity,
                    'unit_price': float(item.unit_price),
                    'total_value': float(item.unit_price) * total_quantity,
                    'locations': locations,
                })
            return Response(report_data, status=status.HTTP_200_OK)
        except Exception as e:
            import traceback
            print("CATEGORIES REPORT ERROR:", e)
            traceback.print_exc()
            return Response(
                {'error': f'Failed to generate categories report: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def generate_register_report(self, request):
        """Generate register report showing stock movements in/out of main store for selected category"""
        try:
            filters = request.data.get('filters', {})
            
            # Validate required filters
            if not filters.get('category'):
                return Response(
                    {'error': 'Category filter is required for register report'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Import time at the top level
            from datetime import time
            
            # Get main store location
            from ..models import Location, ProcurementItem
            try:
                main_store = Location.get_main_store()
                if not main_store:
                    return Response(
                        {'error': 'Main store location not found'}, 
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
            except Exception as e:
                return Response(
                    {'error': f'Failed to get main store location: {str(e)}'}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
            
            # Collect all data entries (both stock movements and procurements)
            all_entries = []
            
            # Parse date filters once at the beginning
            start_date = self._parse_date_filter(filters.get('startDate'))
            end_date = self._parse_date_filter(filters.get('endDate'))
            
            # 1. Get StockMovement records for items in the selected category
            try:
                stock_movements = StockMovement.objects.filter(
                    item__category__id=filters.get('category')
                ).order_by('movement_date', 'id')
                
                # Apply date filters to stock movements
                if start_date:
                    stock_movements = stock_movements.filter(movement_date__gte=start_date.date())
                if end_date:
                    stock_movements = stock_movements.filter(movement_date__lte=end_date.date())
                
                # Filter for movements involving main store (either from or to)
                stock_movements = stock_movements.filter(
                    Q(from_location=main_store) | Q(to_location=main_store)
                )
                
                # Convert stock movements to register entries
                for movement in stock_movements:
                    # Determine if this is a receipt or issue
                    if movement.to_location == main_store:
                        # Receipt into main store
                        received_issued = 'Received'
                        received = movement.quantity
                        issued = ''
                        balance_change = movement.quantity  # Positive for received
                        # Create descriptive remarks for received items
                        remarks = movement.notes or f"Stock movement from {movement.from_location.name} to {movement.to_location.name}"
                    elif movement.from_location == main_store:
                        # Issue from main store
                        received_issued = 'Issued'
                        received = ''
                        issued = movement.quantity
                        balance_change = -movement.quantity  # Negative for issued
                        # Create descriptive remarks for issued items
                        remarks = movement.notes or f"Stock movement from {movement.from_location.name} to {movement.to_location.name}"
                    else:
                        continue
                    
                    # Get voucher number from movement notes or create one
                    voucher_no = movement.notes or f"MOV-{movement.id:06d}"
                    
                    # Create datetime for proper ordering - use morning time for received, afternoon for issued
                    if received_issued == 'Received':
                        # Received entries get morning time (9 AM)
                        movement_time = time(hour=9, minute=0, second=0)
                    else:
                        # Issued entries get afternoon time (2 PM) to ensure they come after received entries
                        movement_time = time(hour=14, minute=0, second=0)
                    
                    movement_datetime = timezone.make_aware(
                        datetime.combine(movement.movement_date, movement_time),
                        timezone=timezone.get_current_timezone()
                    )
                    
                    all_entries.append({
                        'datetime': movement_datetime,  # For sorting
                        'date': movement.movement_date.strftime('%Y-%m-%d'),
                        'receivedIssued': received_issued,
                        'voucherNo': voucher_no,
                        'particulars': movement.item.name,
                        'unit': getattr(movement.item, 'unit', 'PCS'),
                        'unitPrice': float(movement.item.unit_price),
                        'totalCost': float(movement.item.unit_price) * movement.quantity,
                        'quantity': {
                            'received': received if received else '',
                            'issued': issued if issued else '',
                            'balance': 0,  # Will be calculated later
                        },
                        'remarks': remarks,
                        'item_id': movement.item_id,
                        'balance_change': balance_change,  # For balance calculation
                    })
            except Exception as e:
                print(f"Error processing stock movements: {str(e)}")
                # Continue with empty stock movements if there's an error
                pass
            
            # 2. Get Procurement records for items in the selected category
            try:
                procurements = Procurement.objects.filter(
                    items__item__category__id=filters.get('category')
                ).order_by('order_date', 'id')
                
                # Apply date filters to procurements (use order_date for consistency)
                if start_date:
                    procurements = procurements.filter(order_date__gte=start_date.date())
                if end_date:
                    procurements = procurements.filter(order_date__lte=end_date.date())
                
                # Convert procurements to register entries
                for procurement in procurements:
                    for proc_item in procurement.items.all():
                        if proc_item.item.category.id == int(filters.get('category')):
                            # This is a receipt into main store from procurement
                            voucher_no = f"PO-{procurement.order_number}"
                            
                            # Use morning time (8 AM) for procurement entries to ensure they come first
                            procurement_time = timezone.make_aware(
                                datetime.combine(procurement.order_date, time(hour=8, minute=0, second=0)),
                                timezone=timezone.get_current_timezone()
                            )
                            
                            all_entries.append({
                                'datetime': procurement_time,  # For chronological ordering
                                'date': procurement.order_date.strftime('%Y-%m-%d'),
                                'receivedIssued': 'Received',
                                'voucherNo': voucher_no,
                                'particulars': proc_item.item.name,
                                'unit': getattr(proc_item.item, 'unit', 'PCS'),
                                'unitPrice': float(proc_item.unit_price),
                                'totalCost': float(proc_item.unit_price) * proc_item.quantity,
                                'quantity': {
                                    'received': proc_item.quantity,
                                    'issued': '',
                                    'balance': 0,  # Will be calculated later
                                },
                                'remarks': f"Procurement from {procurement.supplier}",
                                'item_id': proc_item.item_id,
                                'balance_change': proc_item.quantity,  # Positive for received
                            })
            except Exception as e:
                print(f"Error processing procurements: {str(e)}")
                # Continue with empty procurements if there's an error
                pass
            
            # Check if we have any data
            if not all_entries:
                # Create a Report object for export functionality even with no data
                report = Report.objects.create(
                    report_type='register',
                    filters=filters,
                    generated_by=self._get_generated_by_user(request)
                )
                
                return Response({
                    'report_id': report.id,
                    'report_type': 'register',
                    'generated_at': report.generated_at,
                    'filters': filters,
                    'total_records': 0,
                    'data': [],
                    'message': 'No data found for the selected category and date range'
                }, status=status.HTTP_200_OK)
            
            # Sort all entries by exact datetime for proper chronological order
            all_entries.sort(key=lambda x: (x['datetime'], x['receivedIssued'] == 'Issued'))
            
            # Calculate running balance per item
            balance_map = {}  # {item_id: balance}
            data = []
            
            for entry in all_entries:
                try:
                    item_id = entry['item_id']
                    prev_balance = balance_map.get(item_id, 0)
                    
                    # Calculate the new balance
                    new_balance = prev_balance + entry['balance_change']
                    balance_map[item_id] = new_balance
                    
                    # Update the balance in the entry
                    entry['quantity']['balance'] = new_balance
                    
                    # Remove helper fields
                    del entry['datetime']
                    del entry['item_id']
                    del entry['balance_change']
                    
                    data.append(entry)
                except Exception as e:
                    print(f"Error processing entry: {str(e)}")
                    # Skip this entry if there's an error
                    continue
            
            # Create a Report object for export functionality
            report = Report.objects.create(
                report_type='register',
                filters=filters,
                generated_by=self._get_generated_by_user(request)
            )
            
            return Response({
                'report_id': report.id,
                'report_type': 'register',
                'generated_at': report.generated_at,
                'filters': filters,
                'total_records': len(data),
                'data': data
            })
            
        except Exception as e:
            import traceback
            print("REGISTER REPORT ERROR:", e)
            traceback.print_exc()
            return Response(
                {'error': f'Failed to generate register report: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            ) 