from rest_framework import generics, filters, views, response
from ..models import AuditLog, User
from ..serializers import AuditLogSerializer
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
import csv
from io import StringIO, BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import os
from django.db import models
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors
from django.conf import settings

class AuditLogListView(generics.ListAPIView):
    queryset = AuditLog.objects.all().order_by('-timestamp')
    serializer_class = AuditLogSerializer
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['action', 'entity_type', 'details', 'performed_by__name']
    filterset_fields = {
        'performed_by': ['exact'],
        'timestamp': ['date__gte', 'date__lte'],
        'action': ['exact'],
        'entity_type': ['exact'],
    }

class AuditLogActionsView(views.APIView):
    def get(self, request):
        actions = AuditLog.objects.values_list('action', flat=True).distinct()
        return response.Response(actions)

class AuditLogEntitiesView(views.APIView):
    def get(self, request):
        entities = AuditLog.objects.values_list('entity_type', flat=True).distinct()
        return response.Response(entities)

class AuditLogUsersView(views.APIView):
    def get(self, request):
        users = User.objects.values('id', 'name')
        return response.Response(list(users))

# Remove the AuditLogExportCSVView class and its logic

class AuditLogExportPDFView(views.APIView):
    def get(self, request):
        import datetime
        logs = AuditLog.objects.all()
        action = request.GET.get('action')
        entity_type = request.GET.get('entity_type')
        performed_by = request.GET.get('performed_by')
        search = request.GET.get('search')
        if action:
            logs = logs.filter(action=action)
        if entity_type:
            logs = logs.filter(entity_type=entity_type)
        if performed_by:
            logs = logs.filter(performed_by_id=performed_by)
        if search:
            logs = logs.filter(
                models.Q(action__icontains=search) |
                models.Q(entity_type__icontains=search) |
                models.Q(details__icontains=search) |
                models.Q(performed_by__name__icontains=search)
            )
        # Prepare data for PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=0.75*inch,
            rightMargin=0.75*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        elements = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#2E5984')
        ))
        styles.add(ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#1E3F66')
        ))
        styles.add(ParagraphStyle(
            'Metadata',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            fontName='Helvetica'
        ))
        styles.add(ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            spaceBefore=20,
            alignment=TA_CENTER,
            textColor=colors.grey
        ))
        # Add logo
        try:
            logo_paths = [
                os.path.join(os.path.dirname(__file__), "nedlogo.jpeg"),
                os.path.join(settings.BASE_DIR, "inventory", "views", "nedlogo.jpeg"),
                "nedlogo.jpeg",
                "inventory/views/nedlogo.jpeg"
            ]
            logo_found = False
            for logo_path in logo_paths:
                if os.path.exists(logo_path):
                    logo = Image(logo_path, width=1*inch, height=1*inch)
                    elements.append(logo)
                    elements.append(Spacer(1, 10))
                    logo_found = True
                    break
            if not logo_found:
                header_text = Paragraph("CSIT Inventory Management System", styles['CustomTitle'])
                elements.append(header_text)
                elements.append(Spacer(1, 10))
        except Exception:
            header_text = Paragraph("CSIT Inventory Management System", styles['CustomTitle'])
            elements.append(header_text)
            elements.append(Spacer(1, 10))
        # Title
        title = Paragraph("Audit Log Report", styles['CustomTitle'])
        elements.append(title)
        elements.append(Spacer(1, 0.1*inch))
        elements.append(HRFlowable(
            width="100%",
            thickness=1,
            lineCap='round',
            color=colors.HexColor('#2E5984'),
            spaceAfter=0.3*inch
        ))
        # Metadata
        metadata = [
            ["Generated on:", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Total Records:", str(logs.count())],
        ]
        metadata_table = Table(metadata, colWidths=[2*inch, 3*inch])
        metadata_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1E3F66')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(metadata_table)
        elements.append(Spacer(1, 0.3*inch))
        # Table
        if logs.exists():
            elements.append(Spacer(1, 0.2*inch))  # More space before table
            elements.append(Paragraph("Report Details", styles['SectionHeader']))
            headers = ['ID', 'Action', 'Entity Type', 'Performed By', 'Timestamp', 'Details']
            table_data = [
                [Paragraph(h, styles['Normal']) for h in headers]
            ]
            for log in logs:
                row = [
                    Paragraph(str(log.id), styles['Normal']),
                    Paragraph(log.action, styles['Normal']),
                    Paragraph(log.entity_type, styles['Normal']),
                    Paragraph(log.performed_by.name if log.performed_by else '', styles['Normal']),
                    Paragraph(log.timestamp.strftime('%Y-%m-%d %H:%M:%S'), styles['Normal']),
                    Paragraph(log.details, styles['Normal']) if log.details else Paragraph('', styles['Normal'])
                ]
                table_data.append(row)
            # Fixed column widths in points (1 inch = 72 points)
            col_widths = [30, 90, 80, 80, 100, 160]
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dce6f1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]
            table.setStyle(TableStyle(table_style))
            elements.append(table)
            elements.append(Spacer(1, 0.2*inch))  # More space after table
        # Footer
        elements.append(Spacer(1, 0.3*inch))
        footer = Paragraph("Confidential - 2025 © NED UET", styles['Footer'])
        elements.append(footer)
        doc.build(elements)
        buffer.seek(0)
        response_pdf = HttpResponse(buffer, content_type='application/pdf')
        response_pdf['Content-Disposition'] = 'attachment; filename="audit_logs.pdf"'
        return response_pdf

class AuditLogExportExcelView(views.APIView):
    def get(self, request):
        import datetime
        logs = AuditLog.objects.all()
        action = request.GET.get('action')
        entity_type = request.GET.get('entity_type')
        performed_by = request.GET.get('performed_by')
        search = request.GET.get('search')
        if action:
            logs = logs.filter(action=action)
        if entity_type:
            logs = logs.filter(entity_type=entity_type)
        if performed_by:
            logs = logs.filter(performed_by_id=performed_by)
        if search:
            logs = logs.filter(
                models.Q(action__icontains=search) |
                models.Q(entity_type__icontains=search) |
                models.Q(details__icontains=search) |
                models.Q(performed_by__name__icontains=search)
            )
        # Prepare Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Log Report"
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        data_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        # Heading
        ws['A1'] = "Audit Log Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal="center")
        # Metadata
        ws['A3'] = f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Total Records: {logs.count()}"
        # Table headers
        headers = ['ID', 'Action', 'Entity Type', 'Performed By', 'Timestamp', 'Details']
        ws.append([])  # Empty row for spacing
        ws.append(headers)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        # Data rows
        for log in logs:
            row = [
                log.id,
                log.action,
                log.entity_type,
                log.performed_by.name if log.performed_by else '',
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                log.details
            ]
            ws.append(row)
        # Style data rows
        for row in ws.iter_rows(min_row=7, max_row=6+logs.count(), min_col=1, max_col=6):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = border
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="audit_logs.xlsx"'
        return response 